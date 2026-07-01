"""Create a SQLite database from Summary_Reviewed_List_to_screen.xlsx.

Description:
Read the screening workbook created by ``04_Get_Repositories_to_screen.py``
and seed a normalized SQLite database for repository analysis. The script
creates the following tables:

- ``repositories``: stable repository metadata and context fields
- ``screening_snapshot``: the initial rolling 14-day screening values
- ``daily_traffic``: daily fact table for later scheduled updates
- ``import_runs``: audit trail of database imports

Parameters:
- ``screening_workbook``: path to ``Summary_Reviewed_List_to_screen.xlsx`` or a
  compatible screening workbook.
- ``output_database``: target SQLite database file. Default:
  ``<repo_root>/data/db/Repository_Analysis.sqlite``.
- ``--replace``: recreate the output database file from scratch before
  importing the workbook.

Calling format:
    python scripts/05_Create_SQLite_Database.py <screening_workbook> [output_database]
    python scripts/05_Create_SQLite_Database.py <screening_workbook> [output_database] --replace

Example:
    python scripts/05_Create_SQLite_Database.py data/screening/Summary_Reviewed_List_to_screen.xlsx --replace

Dependencies:
The script requires ``openpyxl``.
Install it with ``pip install openpyxl`` if it is not already available.
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - handled at runtime
    raise SystemExit(
        "openpyxl is required to read the screening workbook. "
        "Install it with: pip install openpyxl"
    ) from exc


SCRIPT_VERSION = "1.0.0"
DEFAULT_DATABASE_NAME = "Repository_Analysis.sqlite"
DEFAULT_DATABASE_SUBDIR = Path("data") / "db"
DATA_SHEET_NAME = "repositories_to_screen"

EXPECTED_HEADER = [
    "repository",
    "last_push",
    "unique_visitors_14d",
    "unique_cloners_14d",
    "total_views_14d",
    "snapshot_date",
    "organization",
    "repository_full_name",
    "link_2_repository",
    "traffic_status",
    "metric_date",
    "views_count",
    "views_uniques",
    "clones_count",
    "clones_uniques",
    "collected_at",
    "visibility",
    "archived",
    "days_since_last_push",
    "forks_count",
    "stargazers_count",
    "open_issues_count",
]

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS repositories (
    repository_id INTEGER PRIMARY KEY AUTOINCREMENT,
    repository_full_name TEXT NOT NULL UNIQUE,
    repository TEXT NOT NULL,
    organization TEXT NOT NULL,
    link_2_repository TEXT NOT NULL,
    visibility TEXT,
    archived INTEGER CHECK (archived IN (0, 1) OR archived IS NULL),
    last_push TEXT,
    days_since_last_push INTEGER,
    traffic_status TEXT,
    forks_count INTEGER,
    stargazers_count INTEGER,
    open_issues_count INTEGER,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS screening_snapshot (
    screening_snapshot_id INTEGER PRIMARY KEY AUTOINCREMENT,
    repository_id INTEGER NOT NULL,
    snapshot_date TEXT NOT NULL,
    unique_visitors_14d INTEGER,
    unique_cloners_14d INTEGER,
    total_views_14d INTEGER,
    collected_at TEXT,
    source_workbook TEXT NOT NULL,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    UNIQUE(repository_id, snapshot_date),
    FOREIGN KEY (repository_id) REFERENCES repositories(repository_id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS daily_traffic (
    daily_traffic_id INTEGER PRIMARY KEY AUTOINCREMENT,
    repository_id INTEGER NOT NULL,
    metric_date TEXT NOT NULL,
    views_count INTEGER,
    views_uniques INTEGER,
    clones_count INTEGER,
    clones_uniques INTEGER,
    collected_at TEXT,
    source_workbook TEXT NOT NULL,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    UNIQUE(repository_id, metric_date),
    FOREIGN KEY (repository_id) REFERENCES repositories(repository_id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS import_runs (
    import_run_id INTEGER PRIMARY KEY AUTOINCREMENT,
    imported_at TEXT NOT NULL,
    source_workbook TEXT NOT NULL,
    source_script TEXT NOT NULL,
    output_database TEXT NOT NULL,
    rows_read INTEGER NOT NULL,
    repositories_processed INTEGER NOT NULL,
    screening_snapshots_processed INTEGER NOT NULL,
    daily_traffic_processed INTEGER NOT NULL,
    replace_mode INTEGER NOT NULL CHECK (replace_mode IN (0, 1)),
    notes TEXT
);

CREATE INDEX IF NOT EXISTS idx_repositories_organization
    ON repositories (organization);

CREATE INDEX IF NOT EXISTS idx_screening_snapshot_snapshot_date
    ON screening_snapshot (snapshot_date);

CREATE INDEX IF NOT EXISTS idx_daily_traffic_metric_date
    ON daily_traffic (metric_date);
"""


@dataclass(frozen=True)
class ScreeningWorkbookRow:
    repository: str
    last_push: str | None
    unique_visitors_14d: int | None
    unique_cloners_14d: int | None
    total_views_14d: int | None
    snapshot_date: str
    organization: str
    repository_full_name: str
    link_2_repository: str
    traffic_status: str | None
    metric_date: str | None
    views_count: int | None
    views_uniques: int | None
    clones_count: int | None
    clones_uniques: int | None
    collected_at: str | None
    visibility: str | None
    archived: int | None
    days_since_last_push: int | None
    forks_count: int | None
    stargazers_count: int | None
    open_issues_count: int | None


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read Summary_Reviewed_List_to_screen.xlsx and seed a normalized "
            "SQLite database for repository analysis."
        )
    )
    parser.add_argument(
        "screening_workbook",
        type=Path,
        help="Screening workbook path.",
    )
    parser.add_argument(
        "output_database",
        nargs="?",
        type=Path,
        help="Target SQLite database path. Default: <screening_workbook_dir>/Repository_Analysis.sqlite.",
    )
    parser.add_argument(
        "--replace",
        action="store_true",
        help="Recreate the output database file from scratch before importing.",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s {SCRIPT_VERSION}",
    )
    return parser.parse_args(argv)


def resolve_output_database(
    screening_workbook: Path,
    output_database: Path | None,
) -> Path:
    if output_database is not None:
        return output_database.expanduser().resolve()
    return get_repository_root() / DEFAULT_DATABASE_SUBDIR / DEFAULT_DATABASE_NAME


def get_repository_root() -> Path:
    return Path(__file__).resolve().parent.parent


def normalize_row_values(values: list[Any]) -> list[str]:
    normalized: list[str] = []
    for value in values[: len(EXPECTED_HEADER)]:
        normalized.append("" if value is None else str(value).strip())
    return normalized


def find_header_row(worksheet) -> tuple[int, dict[str, int]]:
    for row_index in range(1, worksheet.max_row + 1):
        row_values = [
            worksheet.cell(row=row_index, column=column_index).value
            for column_index in range(1, len(EXPECTED_HEADER) + 1)
        ]
        if normalize_row_values(row_values) == EXPECTED_HEADER:
            return row_index, {name: index + 1 for index, name in enumerate(EXPECTED_HEADER)}

    raise ValueError(
        f"Could not locate the screening data header row in worksheet '{worksheet.title}'."
    )


def cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def normalize_datetime_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=UTC)
        else:
            value = value.astimezone(UTC)
        return value.strftime("%Y-%m-%dT%H:%M:%SZ")
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day, tzinfo=UTC).strftime(
            "%Y-%m-%dT%H:%M:%SZ"
        )

    normalized = str(value).strip()
    if not normalized:
        return None

    try:
        parsed = datetime.fromisoformat(normalized.replace("Z", "+00:00"))
    except ValueError:
        try:
            parsed_date = datetime.strptime(normalized, "%Y-%m-%d").date()
        except ValueError:
            return normalized
        return datetime(
            parsed_date.year,
            parsed_date.month,
            parsed_date.day,
            tzinfo=UTC,
        ).strftime("%Y-%m-%dT%H:%M:%SZ")

    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=UTC)
    else:
        parsed = parsed.astimezone(UTC)
    return parsed.strftime("%Y-%m-%dT%H:%M:%SZ")


def normalize_date_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    normalized = str(value).strip()
    if not normalized:
        return None

    try:
        return datetime.fromisoformat(normalized.replace("Z", "+00:00")).date().isoformat()
    except ValueError:
        try:
            return datetime.strptime(normalized, "%Y-%m-%d").date().isoformat()
        except ValueError:
            return normalized


def normalize_integer(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)

    normalized = str(value).strip()
    if not normalized:
        return None

    try:
        return int(float(normalized))
    except ValueError:
        return None


def normalize_boolean_integer(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return 1 if value else 0

    normalized = str(value).strip().upper()
    if not normalized:
        return None
    if normalized == "TRUE":
        return 1
    if normalized == "FALSE":
        return 0
    return None


def extract_repository_link(link_cell) -> str:
    if link_cell.hyperlink and link_cell.hyperlink.target:
        return link_cell.hyperlink.target.strip()
    return cell_to_string(link_cell.value)


def derive_repository_name(repository: str, repository_full_name: str) -> str:
    if repository:
        return repository
    if "/" in repository_full_name:
        return repository_full_name.split("/", 1)[1]
    return repository_full_name


def derive_organization(organization: str, repository_full_name: str) -> str:
    if organization:
        return organization
    if "/" in repository_full_name:
        return repository_full_name.split("/", 1)[0]
    return ""


def read_screening_rows(screening_workbook: Path) -> list[ScreeningWorkbookRow]:
    try:
        workbook = load_workbook(screening_workbook, data_only=False)
    except PermissionError as error:
        raise SystemExit(
            f"Cannot open workbook '{screening_workbook}'. Close the file in Excel and retry."
        ) from error

    if DATA_SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(
            f"Worksheet '{DATA_SHEET_NAME}' was not found in {screening_workbook}."
        )

    worksheet = workbook[DATA_SHEET_NAME]
    header_row, column_map = find_header_row(worksheet)
    selected_rows: list[ScreeningWorkbookRow] = []

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        repository_full_name = cell_to_string(
            worksheet.cell(
                row=row_index,
                column=column_map["repository_full_name"],
            ).value
        )
        repository = cell_to_string(
            worksheet.cell(row=row_index, column=column_map["repository"]).value
        )
        organization = cell_to_string(
            worksheet.cell(row=row_index, column=column_map["organization"]).value
        )
        if not repository and not repository_full_name:
            continue

        repository_full_name = repository_full_name or (
            f"{organization}/{repository}" if organization and repository else ""
        )
        if not repository_full_name:
            raise ValueError(
                f"Missing repository_full_name for worksheet row {row_index} in {screening_workbook}."
            )

        snapshot_date = normalize_date_text(
            worksheet.cell(row=row_index, column=column_map["snapshot_date"]).value
        )
        if not snapshot_date:
            raise ValueError(
                f"Missing snapshot_date for repository '{repository_full_name}'."
            )

        selected_rows.append(
            ScreeningWorkbookRow(
                repository=derive_repository_name(repository, repository_full_name),
                last_push=normalize_datetime_text(
                    worksheet.cell(row=row_index, column=column_map["last_push"]).value
                ),
                unique_visitors_14d=normalize_integer(
                    worksheet.cell(
                        row=row_index,
                        column=column_map["unique_visitors_14d"],
                    ).value
                ),
                unique_cloners_14d=normalize_integer(
                    worksheet.cell(
                        row=row_index,
                        column=column_map["unique_cloners_14d"],
                    ).value
                ),
                total_views_14d=normalize_integer(
                    worksheet.cell(
                        row=row_index,
                        column=column_map["total_views_14d"],
                    ).value
                ),
                snapshot_date=snapshot_date,
                organization=derive_organization(organization, repository_full_name),
                repository_full_name=repository_full_name,
                link_2_repository=extract_repository_link(
                    worksheet.cell(row=row_index, column=column_map["link_2_repository"])
                ),
                traffic_status=cell_to_string(
                    worksheet.cell(
                        row=row_index,
                        column=column_map["traffic_status"],
                    ).value
                )
                or None,
                metric_date=normalize_date_text(
                    worksheet.cell(row=row_index, column=column_map["metric_date"]).value
                ),
                views_count=normalize_integer(
                    worksheet.cell(row=row_index, column=column_map["views_count"]).value
                ),
                views_uniques=normalize_integer(
                    worksheet.cell(row=row_index, column=column_map["views_uniques"]).value
                ),
                clones_count=normalize_integer(
                    worksheet.cell(row=row_index, column=column_map["clones_count"]).value
                ),
                clones_uniques=normalize_integer(
                    worksheet.cell(
                        row=row_index,
                        column=column_map["clones_uniques"],
                    ).value
                ),
                collected_at=normalize_datetime_text(
                    worksheet.cell(row=row_index, column=column_map["collected_at"]).value
                ),
                visibility=cell_to_string(
                    worksheet.cell(row=row_index, column=column_map["visibility"]).value
                )
                or None,
                archived=normalize_boolean_integer(
                    worksheet.cell(row=row_index, column=column_map["archived"]).value
                ),
                days_since_last_push=normalize_integer(
                    worksheet.cell(
                        row=row_index,
                        column=column_map["days_since_last_push"],
                    ).value
                ),
                forks_count=normalize_integer(
                    worksheet.cell(row=row_index, column=column_map["forks_count"]).value
                ),
                stargazers_count=normalize_integer(
                    worksheet.cell(
                        row=row_index,
                        column=column_map["stargazers_count"],
                    ).value
                ),
                open_issues_count=normalize_integer(
                    worksheet.cell(
                        row=row_index,
                        column=column_map["open_issues_count"],
                    ).value
                ),
            )
        )

    return selected_rows


def ensure_database_target(output_database: Path, replace: bool) -> None:
    output_database.parent.mkdir(parents=True, exist_ok=True)
    if replace and output_database.exists():
        output_database.unlink()


def create_schema(connection: sqlite3.Connection) -> None:
    connection.executescript(SCHEMA_SQL)


def upsert_repository(
    connection: sqlite3.Connection,
    row: ScreeningWorkbookRow,
    imported_at: str,
) -> int:
    connection.execute(
        """
        INSERT INTO repositories (
            repository_full_name,
            repository,
            organization,
            link_2_repository,
            visibility,
            archived,
            last_push,
            days_since_last_push,
            traffic_status,
            forks_count,
            stargazers_count,
            open_issues_count,
            created_at,
            updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(repository_full_name) DO UPDATE SET
            repository = excluded.repository,
            organization = excluded.organization,
            link_2_repository = excluded.link_2_repository,
            visibility = COALESCE(excluded.visibility, repositories.visibility),
            archived = CASE
                WHEN excluded.archived IS NOT NULL THEN excluded.archived
                ELSE repositories.archived
            END,
            last_push = COALESCE(excluded.last_push, repositories.last_push),
            days_since_last_push = CASE
                WHEN excluded.days_since_last_push IS NOT NULL THEN excluded.days_since_last_push
                ELSE repositories.days_since_last_push
            END,
            traffic_status = COALESCE(excluded.traffic_status, repositories.traffic_status),
            forks_count = CASE
                WHEN excluded.forks_count IS NOT NULL THEN excluded.forks_count
                ELSE repositories.forks_count
            END,
            stargazers_count = CASE
                WHEN excluded.stargazers_count IS NOT NULL THEN excluded.stargazers_count
                ELSE repositories.stargazers_count
            END,
            open_issues_count = CASE
                WHEN excluded.open_issues_count IS NOT NULL THEN excluded.open_issues_count
                ELSE repositories.open_issues_count
            END,
            updated_at = excluded.updated_at
        """,
        (
            row.repository_full_name,
            row.repository,
            row.organization,
            row.link_2_repository,
            row.visibility,
            row.archived,
            row.last_push,
            row.days_since_last_push,
            row.traffic_status,
            row.forks_count,
            row.stargazers_count,
            row.open_issues_count,
            imported_at,
            imported_at,
        ),
    )

    result = connection.execute(
        "SELECT repository_id FROM repositories WHERE repository_full_name = ?",
        (row.repository_full_name,),
    ).fetchone()
    if result is None:
        raise RuntimeError(f"Failed to resolve repository_id for {row.repository_full_name}.")
    return int(result[0])


def upsert_screening_snapshot(
    connection: sqlite3.Connection,
    repository_id: int,
    row: ScreeningWorkbookRow,
    screening_workbook: Path,
    imported_at: str,
) -> None:
    connection.execute(
        """
        INSERT INTO screening_snapshot (
            repository_id,
            snapshot_date,
            unique_visitors_14d,
            unique_cloners_14d,
            total_views_14d,
            collected_at,
            source_workbook,
            created_at,
            updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(repository_id, snapshot_date) DO UPDATE SET
            unique_visitors_14d = excluded.unique_visitors_14d,
            unique_cloners_14d = excluded.unique_cloners_14d,
            total_views_14d = excluded.total_views_14d,
            collected_at = COALESCE(excluded.collected_at, screening_snapshot.collected_at),
            source_workbook = excluded.source_workbook,
            updated_at = excluded.updated_at
        """,
        (
            repository_id,
            row.snapshot_date,
            row.unique_visitors_14d,
            row.unique_cloners_14d,
            row.total_views_14d,
            row.collected_at,
            str(screening_workbook),
            imported_at,
            imported_at,
        ),
    )


def upsert_daily_traffic(
    connection: sqlite3.Connection,
    repository_id: int,
    row: ScreeningWorkbookRow,
    screening_workbook: Path,
    imported_at: str,
) -> None:
    connection.execute(
        """
        INSERT INTO daily_traffic (
            repository_id,
            metric_date,
            views_count,
            views_uniques,
            clones_count,
            clones_uniques,
            collected_at,
            source_workbook,
            created_at,
            updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(repository_id, metric_date) DO UPDATE SET
            views_count = excluded.views_count,
            views_uniques = excluded.views_uniques,
            clones_count = excluded.clones_count,
            clones_uniques = excluded.clones_uniques,
            collected_at = COALESCE(excluded.collected_at, daily_traffic.collected_at),
            source_workbook = excluded.source_workbook,
            updated_at = excluded.updated_at
        """,
        (
            repository_id,
            row.metric_date,
            row.views_count,
            row.views_uniques,
            row.clones_count,
            row.clones_uniques,
            row.collected_at,
            str(screening_workbook),
            imported_at,
            imported_at,
        ),
    )


def record_import_run(
    connection: sqlite3.Connection,
    imported_at: str,
    screening_workbook: Path,
    output_database: Path,
    rows_read: int,
    repositories_processed: int,
    screening_snapshots_processed: int,
    daily_traffic_processed: int,
    replace_mode: bool,
) -> None:
    connection.execute(
        """
        INSERT INTO import_runs (
            imported_at,
            source_workbook,
            source_script,
            output_database,
            rows_read,
            repositories_processed,
            screening_snapshots_processed,
            daily_traffic_processed,
            replace_mode,
            notes
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            imported_at,
            str(screening_workbook),
            f"05_Create_SQLite_Database.py v{SCRIPT_VERSION}",
            str(output_database),
            rows_read,
            repositories_processed,
            screening_snapshots_processed,
            daily_traffic_processed,
            1 if replace_mode else 0,
            (
                "Daily traffic rows are imported only when metric_date is populated "
                "in the screening workbook."
            ),
        ),
    )


def import_workbook_to_database(
    screening_workbook: Path,
    output_database: Path,
    rows: list[ScreeningWorkbookRow],
    replace_mode: bool,
) -> tuple[int, int, int]:
    imported_at = datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")

    with sqlite3.connect(output_database) as connection:
        connection.execute("PRAGMA foreign_keys = ON")
        create_schema(connection)

        repositories_processed = 0
        screening_snapshots_processed = 0
        daily_traffic_processed = 0

        for row in rows:
            repository_id = upsert_repository(connection, row, imported_at)
            repositories_processed += 1

            upsert_screening_snapshot(
                connection,
                repository_id,
                row,
                screening_workbook,
                imported_at,
            )
            screening_snapshots_processed += 1

            if row.metric_date:
                upsert_daily_traffic(
                    connection,
                    repository_id,
                    row,
                    screening_workbook,
                    imported_at,
                )
                daily_traffic_processed += 1

        record_import_run(
            connection,
            imported_at,
            screening_workbook,
            output_database,
            len(rows),
            repositories_processed,
            screening_snapshots_processed,
            daily_traffic_processed,
            replace_mode,
        )

    return repositories_processed, screening_snapshots_processed, daily_traffic_processed


def print_summary(
    screening_workbook: Path,
    output_database: Path,
    rows_read: int,
    repositories_processed: int,
    screening_snapshots_processed: int,
    daily_traffic_processed: int,
) -> None:
    print(f"Screening workbook: {screening_workbook}")
    print(f"Rows read: {rows_read}")
    print(f"Repositories processed: {repositories_processed}")
    print(f"Screening snapshots processed: {screening_snapshots_processed}")
    print(f"Daily traffic rows processed: {daily_traffic_processed}")
    print(f"Output database: {output_database}")


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    screening_workbook = args.screening_workbook.expanduser().resolve()
    if not screening_workbook.is_file():
        raise SystemExit(f"Screening workbook not found: {screening_workbook}")

    output_database = resolve_output_database(
        screening_workbook,
        args.output_database,
    )
    ensure_database_target(output_database, args.replace)

    rows = read_screening_rows(screening_workbook)
    if not rows:
        raise SystemExit(f"No repository rows were found in {screening_workbook}.")

    repositories_processed, screening_snapshots_processed, daily_traffic_processed = (
        import_workbook_to_database(
            screening_workbook,
            output_database,
            rows,
            args.replace,
        )
    )
    print_summary(
        screening_workbook,
        output_database,
        len(rows),
        repositories_processed,
        screening_snapshots_processed,
        daily_traffic_processed,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
