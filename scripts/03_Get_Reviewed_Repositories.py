"""Read reviewed repository workbook rows marked USE and validate GitHub access.

Description:
Read the reviewed Excel workbook created from Summary_Raw_List.xlsx, scan all
worksheets, filter rows whose ``ignore_status`` column is set to ``USE``, and
access only those repositories through the GitHub repository API. The selected
rows and the access result are written to a comma-separated output file.

Parameters:
- ``reviewed_workbook``: path to ``Summary_Reviewed_List.xlsx`` or a compatible workbook.
- ``output_file``: target CSV file. Default: ``<reviewed_workbook_stem>_use_access.csv``.

Calling format:
    python scripts/03_Get_Reviewed_Repositories.py <reviewed_workbook> [output_file]

Example:
    python scripts/03_Get_Reviewed_Repositories.py data/reviewed/Summary_Reviewed_List.xlsx

Authentication:
The script uses ``GITHUB_TOKEN`` or ``GH_TOKEN`` when available and otherwise
tries ``gh auth token``. Public repositories can often be read without a token,
but authenticated access is recommended to avoid rate limits.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlparse, urlencode
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - startup guard
    raise SystemExit(
        "Missing dependency: openpyxl. Install it with `pip install openpyxl`."
    ) from exc


SCRIPT_VERSION = "1.0.1"
API_ROOT = "https://api.github.com"
API_VERSION = "2026-03-10"
REQUEST_TIMEOUT_SECONDS = 30
DEFAULT_OUTPUT_SUFFIX = "_use_access.csv"
USE_STATUS = "USE"

EXPECTED_HEADER = [
    "repository",
    "last_push",
    "visibility",
    "archived",
    "unique_visitors_14d",
    "unique_cloners_14d",
    "total_views_14d",
    "traffic_status",
    "link_2_repository",
    "ignore_status",
]

OUTPUT_COLUMNS = [
    "organization",
    "sheet_name",
    "repository",
    "last_push",
    "visibility",
    "archived",
    "unique_visitors_14d",
    "unique_cloners_14d",
    "total_views_14d",
    "traffic_status",
    "link_2_repository",
    "ignore_status",
    "access_status",
    "api_visibility",
    "api_archived",
    "forks_count",
    "default_branch",
    "api_pushed_at",
    "api_html_url",
    "api_full_name",
    "api_status_message",
]

FIELD_DESCRIPTIONS = {
    "organization": "GitHub organization resolved from the reviewed workbook entry.",
    "sheet_name": "Worksheet name in the reviewed workbook.",
    "repository": "Repository name from the reviewed workbook.",
    "last_push": "Last push timestamp copied from the reviewed workbook.",
    "visibility": "Visibility copied from the reviewed workbook.",
    "archived": "Archived flag copied from the reviewed workbook.",
    "unique_visitors_14d": "Unique visitors in the last 14 days copied from the reviewed workbook.",
    "unique_cloners_14d": "Unique cloners in the last 14 days copied from the reviewed workbook.",
    "total_views_14d": "Total views in the last 14 days copied from the reviewed workbook.",
    "traffic_status": "Traffic readout result copied from the reviewed workbook.",
    "link_2_repository": "Repository URL resolved from the workbook hyperlink column.",
    "ignore_status": "Manual review decision copied from the reviewed workbook.",
    "access_status": "GitHub repository access result for rows marked USE.",
    "api_visibility": "Visibility returned by the GitHub repository API.",
    "api_archived": "Archived flag returned by the GitHub repository API.",
    "forks_count": "Fork count returned by the GitHub repository API.",
    "default_branch": "Default branch returned by the GitHub repository API.",
    "api_pushed_at": "Latest push timestamp returned by the GitHub repository API.",
    "api_html_url": "Canonical repository URL returned by the GitHub repository API.",
    "api_full_name": "Full owner/repository name returned by the GitHub repository API.",
    "api_status_message": "Detailed access message returned or derived from the GitHub API response.",
}

SHEET_ORGANIZATION_ALIASES = {
    "open-cmisis-pack_raw_list": "open-cmsis-pack",
    "open-cmisis-pack_reviewed_list": "open-cmsis-pack",
}


@dataclass(frozen=True)
class ReviewedRepositoryRow:
    organization: str
    sheet_name: str
    repository: str
    last_push: str
    visibility: str
    archived: str
    unique_visitors_14d: str
    unique_cloners_14d: str
    total_views_14d: str
    traffic_status: str
    link_2_repository: str
    ignore_status: str


@dataclass(frozen=True)
class RepositoryAccessResult:
    access_status: str
    api_visibility: str
    api_archived: str
    forks_count: str
    default_branch: str
    api_pushed_at: str
    api_html_url: str
    api_full_name: str
    api_status_message: str


class GitHubApiError(RuntimeError):
    """Raised when the GitHub API returns an error response."""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read Summary_Reviewed_List.xlsx, keep only ignore_status=USE rows, "
            "and validate repository access through the GitHub API."
        )
    )
    parser.add_argument("reviewed_workbook", type=Path, help="Reviewed workbook path.")
    parser.add_argument(
        "output_file",
        nargs="?",
        type=Path,
        help="Target CSV output file. Default: <reviewed_workbook_stem>_use_access.csv.",
    )
    return parser.parse_args()


def resolve_output_file(reviewed_workbook: Path, output_file: Path | None) -> Path:
    if output_file is not None:
        return output_file.expanduser().resolve()
    return reviewed_workbook.with_name(reviewed_workbook.stem + DEFAULT_OUTPUT_SUFFIX)


def resolve_token() -> str:
    for env_name in ("GITHUB_TOKEN", "GH_TOKEN"):
        token = os.getenv(env_name, "").strip()
        if token:
            return token

    if shutil.which("gh"):
        completed = subprocess.run(
            ["gh", "auth", "token"],
            check=False,
            capture_output=True,
            text=True,
        )
        token = completed.stdout.strip()
        if completed.returncode == 0 and token:
            return token

    return ""


def build_headers(token: str) -> dict[str, str]:
    headers = {
        "Accept": "application/vnd.github+json",
        "User-Agent": f"03_Get_Reviewed_Repositories.py/{SCRIPT_VERSION}",
        "X-GitHub-Api-Version": API_VERSION,
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return headers


def request_json(token: str, path: str, params: dict[str, Any] | None = None) -> Any:
    query = ""
    if params:
        query = "?" + urlencode(params)

    request = Request(API_ROOT + path + query, headers=build_headers(token))
    try:
        with urlopen(request, timeout=REQUEST_TIMEOUT_SECONDS) as response:
            payload = response.read().decode("utf-8")
            return json.loads(payload)
    except HTTPError as error:
        message = read_error_message(error)
        accepted_permissions = error.headers.get("X-Accepted-GitHub-Permissions", "")
        if accepted_permissions:
            message = f"{message} | accepted_permissions={accepted_permissions}"
        raise GitHubApiError(
            f"GitHub API request failed for {path}: HTTP {error.code} {message}"
        ) from error
    except URLError as error:
        raise GitHubApiError(f"Network error while calling GitHub API: {error}") from error


def read_error_message(error: HTTPError) -> str:
    try:
        payload = error.read().decode("utf-8")
        data = json.loads(payload)
    except Exception:
        return str(error.reason or "Unknown error")

    if isinstance(data, dict):
        message = str(data.get("message", "")).strip()
        errors = data.get("errors")
        if message and errors:
            return f"{message}; errors={errors}"
        if message:
            return message
    return str(error.reason or "Unknown error")


def normalize_row_values(values: list[Any]) -> list[str]:
    normalized: list[str] = []
    for value in values[: len(EXPECTED_HEADER)]:
        normalized.append("" if value is None else str(value).strip())
    return normalized


def find_header_row(worksheet) -> tuple[int, dict[str, int]]:
    for row_index in range(1, worksheet.max_row + 1):
        row_values = [
            worksheet.cell(row=row_index, column=column_index).value
            for column_index in range(1, len(EXPECTED_HEADER) + 1)
        ]
        if normalize_row_values(row_values) == EXPECTED_HEADER:
            return row_index, {name: index + 1 for index, name in enumerate(EXPECTED_HEADER)}

    raise ValueError(
        f"Could not locate the repository data header row in worksheet '{worksheet.title}'."
    )


def cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=UTC)
        else:
            value = value.astimezone(UTC)
        return value.strftime("%Y-%m-%dT%H:%M:%SZ")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value).strip()


def derive_organization_from_sheet(sheet_name: str) -> str:
    if sheet_name in SHEET_ORGANIZATION_ALIASES:
        return SHEET_ORGANIZATION_ALIASES[sheet_name]
    for suffix in ("_raw_list", "_reviewed_list"):
        if sheet_name.endswith(suffix):
            return sheet_name[: -len(suffix)]
    return sheet_name


def extract_repository_url(sheet_name: str, repository: str, link_cell) -> str:
    if link_cell.hyperlink and link_cell.hyperlink.target:
        return link_cell.hyperlink.target.strip()

    value = cell_to_string(link_cell.value)
    if value.startswith("http://") or value.startswith("https://"):
        return value

    if value.startswith("=HYPERLINK("):
        match = re.search(r'"https://github\.com/"\s*&\s*"([^"]+)"', value)
        if match:
            return f"https://github.com/{match.group(1)}/{repository}"

    organization = derive_organization_from_sheet(sheet_name)
    return f"https://github.com/{organization}/{repository}"


def extract_owner_and_repo(repository_url: str) -> tuple[str, str]:
    parsed = urlparse(repository_url)
    parts = [part for part in parsed.path.split("/") if part]
    if len(parts) < 2:
        raise ValueError(f"Unable to parse owner/repository from URL: {repository_url}")
    return parts[0], parts[1]


def read_use_rows(reviewed_workbook: Path) -> list[ReviewedRepositoryRow]:
    try:
        workbook = load_workbook(reviewed_workbook, data_only=False)
    except PermissionError as error:
        raise SystemExit(
            f"Cannot open workbook '{reviewed_workbook}'. Close the file in Excel and retry."
        ) from error

    selected_rows: list[ReviewedRepositoryRow] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row, column_map = find_header_row(worksheet)

        for row_index in range(header_row + 1, worksheet.max_row + 1):
            repository = cell_to_string(
                worksheet.cell(row=row_index, column=column_map["repository"]).value
            )
            if not repository:
                continue

            ignore_status = cell_to_string(
                worksheet.cell(row=row_index, column=column_map["ignore_status"]).value
            ).upper()
            if ignore_status != USE_STATUS:
                continue

            link_cell = worksheet.cell(row=row_index, column=column_map["link_2_repository"])
            repository_url = extract_repository_url(sheet_name, repository, link_cell)
            organization, _ = extract_owner_and_repo(repository_url)

            selected_rows.append(
                ReviewedRepositoryRow(
                    organization=organization,
                    sheet_name=sheet_name,
                    repository=repository,
                    last_push=cell_to_string(
                        worksheet.cell(row=row_index, column=column_map["last_push"]).value
                    ),
                    visibility=cell_to_string(
                        worksheet.cell(row=row_index, column=column_map["visibility"]).value
                    ),
                    archived=cell_to_string(
                        worksheet.cell(row=row_index, column=column_map["archived"]).value
                    ),
                    unique_visitors_14d=cell_to_string(
                        worksheet.cell(
                            row=row_index, column=column_map["unique_visitors_14d"]
                        ).value
                    ),
                    unique_cloners_14d=cell_to_string(
                        worksheet.cell(
                            row=row_index, column=column_map["unique_cloners_14d"]
                        ).value
                    ),
                    total_views_14d=cell_to_string(
                        worksheet.cell(
                            row=row_index, column=column_map["total_views_14d"]
                        ).value
                    ),
                    traffic_status=cell_to_string(
                        worksheet.cell(row=row_index, column=column_map["traffic_status"]).value
                    ),
                    link_2_repository=repository_url,
                    ignore_status=ignore_status,
                )
            )

    return selected_rows


def build_access_status(error: GitHubApiError) -> tuple[str, str]:
    message = str(error)
    if "HTTP 403" in message:
        return "RESTRICTED", message.split("HTTP 403", 1)[-1].strip()
    if "HTTP 404" in message:
        return "NOT_FOUND", message.split("HTTP 404", 1)[-1].strip()
    if "HTTP " in message:
        return "ERROR", message.split("GitHub API request failed for ", 1)[-1]
    return "ERROR", message


def get_repository_access(token: str, organization: str, repository: str) -> RepositoryAccessResult:
    owner = quote(organization, safe="")
    repo = quote(repository, safe="")

    try:
        repo_data = request_json(token, f"/repos/{owner}/{repo}")
    except GitHubApiError as error:
        access_status, message = build_access_status(error)
        return RepositoryAccessResult(
            access_status=access_status,
            api_visibility="",
            api_archived="",
            forks_count="",
            default_branch="",
            api_pushed_at="",
            api_html_url="",
            api_full_name="",
            api_status_message=message,
        )

    return RepositoryAccessResult(
        access_status="OK",
        api_visibility=cell_to_string(repo_data.get("visibility")),
        api_archived=cell_to_string(repo_data.get("archived")),
        forks_count=cell_to_string(repo_data.get("forks_count")),
        default_branch=cell_to_string(repo_data.get("default_branch")),
        api_pushed_at=cell_to_string(parse_timestamp(repo_data.get("pushed_at"))),
        api_html_url=cell_to_string(repo_data.get("html_url")),
        api_full_name=cell_to_string(repo_data.get("full_name")),
        api_status_message="OK",
    )


def parse_timestamp(timestamp: str | None) -> datetime | None:
    if not timestamp:
        return None
    return datetime.strptime(timestamp, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=UTC)


def write_csv(
    output_path: Path,
    reviewed_workbook: Path,
    selected_rows: list[ReviewedRepositoryRow],
    access_results: dict[tuple[str, str], RepositoryAccessResult],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    generated_on = datetime.now(UTC).strftime("%Y-%m-%d")

    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)

        writer.writerow(["# Revision History"])
        writer.writerow(["# Author", "OpenAI Codex"])
        writer.writerow(["# Date", generated_on])
        writer.writerow(
            [
                "# Change",
                (
                    "Selected reviewed workbook rows with ignore_status=USE and "
                    f"validated GitHub repository access using 03_Get_Reviewed_Repositories.py v{SCRIPT_VERSION}"
                ),
            ]
        )
        writer.writerow([])
        writer.writerow(["# File Description"])
        writer.writerow(
            [
                "# Summary",
                (
                    "Repositories were read from the reviewed workbook, filtered by "
                    "ignore_status=USE, and checked through the GitHub repository API."
                ),
            ]
        )
        writer.writerow(
            [
                "# Generation Process",
                (
                    "Two-stage process: stage 1 manual review in Summary_Reviewed_List.xlsx; "
                    "stage 2 automated access validation for the selected repositories."
                ),
            ]
        )
        writer.writerow(["# Source Workbook", str(reviewed_workbook)])
        writer.writerow(
            ["# Source Script", f"03_Get_Reviewed_Repositories.py v{SCRIPT_VERSION}"]
        )
        writer.writerow([])
        writer.writerow(["# Field Descriptions"])
        for field_name, description in FIELD_DESCRIPTIONS.items():
            writer.writerow([f"# {field_name}", description])
        writer.writerow([])
        writer.writerow(OUTPUT_COLUMNS)

        for row in selected_rows:
            result = access_results[(row.organization, row.repository)]
            writer.writerow(
                [
                    row.organization,
                    row.sheet_name,
                    row.repository,
                    row.last_push,
                    row.visibility,
                    row.archived,
                    row.unique_visitors_14d,
                    row.unique_cloners_14d,
                    row.total_views_14d,
                    row.traffic_status,
                    row.link_2_repository,
                    row.ignore_status,
                    result.access_status,
                    result.api_visibility,
                    result.api_archived,
                    result.forks_count,
                    result.default_branch,
                    result.api_pushed_at,
                    result.api_html_url,
                    result.api_full_name,
                    result.api_status_message,
                ]
            )


def print_summary(
    reviewed_workbook: Path,
    output_path: Path,
    selected_rows: list[ReviewedRepositoryRow],
    access_results: dict[tuple[str, str], RepositoryAccessResult],
) -> None:
    ok_count = 0
    failed_count = 0
    for result in access_results.values():
        if result.access_status == "OK":
            ok_count += 1
        else:
            failed_count += 1

    print(f"Reviewed workbook: {reviewed_workbook}")
    print(f"Rows selected with ignore_status={USE_STATUS}: {len(selected_rows)}")
    print(f"Repositories accessed successfully: {ok_count}")
    print(f"Repositories with access issues: {failed_count}")
    print(f"Output file: {output_path}")


def main() -> None:
    args = parse_args()
    reviewed_workbook = args.reviewed_workbook.expanduser().resolve()
    if not reviewed_workbook.is_file():
        raise SystemExit(f"Reviewed workbook not found: {reviewed_workbook}")

    output_path = resolve_output_file(reviewed_workbook, args.output_file)
    token = resolve_token()
    selected_rows = read_use_rows(reviewed_workbook)

    if not selected_rows:
        raise SystemExit(
            f"No rows with ignore_status={USE_STATUS} were found in {reviewed_workbook}."
        )

    access_results: dict[tuple[str, str], RepositoryAccessResult] = {}
    for index, row in enumerate(selected_rows, start=1):
        print(
            f"[{index}/{len(selected_rows)}] Accessing {row.organization}/{row.repository}",
            file=sys.stderr,
        )
        access_results[(row.organization, row.repository)] = get_repository_access(
            token,
            row.organization,
            row.repository,
        )

    write_csv(output_path, reviewed_workbook, selected_rows, access_results)
    print_summary(reviewed_workbook, output_path, selected_rows, access_results)


if __name__ == "__main__":
    main()
