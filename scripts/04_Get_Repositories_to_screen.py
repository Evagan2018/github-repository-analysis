"""Create a screening workbook from Summary_Reviewed_List.xlsx.

Description:
Read the reviewed Excel workbook, scan all worksheets, keep only rows whose
``ignore_status`` column is set to ``USE``, and write a consolidated screening
workbook in ``.xlsx`` format. The output is intended as an editable reference
file and as a clean starting point for later SQLite imports.

Parameters:
- ``reviewed_workbook``: path to ``Summary_Reviewed_List.xlsx`` or a compatible
  reviewed workbook.
- ``output_workbook``: target ``.xlsx`` file. Default:
  ``<repo_root>/data/screening/<reviewed_workbook_stem>_to_screen.xlsx``.
- ``--snapshot-date``: optional snapshot date in ``YYYY-MM-DD`` format.
  Default: today's UTC date.

Calling format:
    python scripts/04_Get_Repositories_to_screen.py <reviewed_workbook> [output_workbook]
    python scripts/04_Get_Repositories_to_screen.py <reviewed_workbook> [output_workbook] --snapshot-date <YYYY-MM-DD>

Example:
    python scripts/04_Get_Repositories_to_screen.py data/reviewed/Summary_Reviewed_List.xlsx

Dependencies:
The script requires ``openpyxl``.
Install it with ``pip install openpyxl`` if it is not already available.
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover - handled at runtime
    raise SystemExit(
        "openpyxl is required to generate the screening workbook. "
        "Install it with: pip install openpyxl"
    ) from exc


SCRIPT_VERSION = "1.0.1"
DEFAULT_OUTPUT_SUFFIX = "_to_screen.xlsx"
DEFAULT_OUTPUT_SUBDIR = Path("data") / "screening"
USE_STATUS = "USE"
DATA_SHEET_NAME = "repositories_to_screen"
README_SHEET_NAME = "README"
DATE_TIME_NUMBER_FORMAT = 'yyyy-mm-dd"T"hh:mm:ss"Z"'
DATE_NUMBER_FORMAT = "yyyy-mm-dd"
INTEGER_NUMBER_FORMAT = "0"
HEADER_FILL = PatternFill(fill_type="solid", fgColor="70AD47")
ALTERNATING_ROW_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
HYPERLINK_FONT = Font(color="0563C1", underline="single")

EXPECTED_HEADER = [
    "repository",
    "last_push",
    "visibility",
    "archived",
    "unique_visitors_14d",
    "unique_cloners_14d",
    "total_views_14d",
    "traffic_status",
    "link_2_repository",
    "ignore_status",
]

OUTPUT_COLUMNS = [
    "repository",
    "last_push",
    "unique_visitors_14d",
    "unique_cloners_14d",
    "total_views_14d",
    "snapshot_date",
    "organization",
    "repository_full_name",
    "link_2_repository",
    "traffic_status",
    "metric_date",
    "views_count",
    "views_uniques",
    "clones_count",
    "clones_uniques",
    "collected_at",
    "visibility",
    "archived",
    "days_since_last_push",
    "forks_count",
    "stargazers_count",
    "open_issues_count",
]

OUTPUT_COLUMN_INDEX = {
    field_name: index + 1 for index, field_name in enumerate(OUTPUT_COLUMNS)
}

FIELD_DESCRIPTIONS = {
    "repository": "Repository name as reviewed in Summary_Reviewed_List.xlsx.",
    "last_push": "Most recent push timestamp copied from the reviewed workbook.",
    "unique_visitors_14d": "Unique visitors in GitHub's rolling last-14-days window. Useful for initial screening only.",
    "unique_cloners_14d": "Unique cloners in GitHub's rolling last-14-days window. Useful for initial screening only.",
    "total_views_14d": "Total repository views in GitHub's rolling last-14-days window. Useful for initial screening only.",
    "snapshot_date": "Date assigned to this screening snapshot. Useful as the time key for SQLite trend tables.",
    "organization": "GitHub organization resolved from the worksheet or repository link.",
    "repository_full_name": "Stable repository key in owner/repository format for joins and trend analysis.",
    "link_2_repository": "Direct clickable repository URL.",
    "traffic_status": "Traffic readout state copied from the reviewed workbook.",
    "metric_date": "Blank placeholder for the GitHub daily traffic bucket date. Populate this in the later daily collector.",
    "views_count": "Blank placeholder for daily total views. Do not copy total_views_14d into this column.",
    "views_uniques": "Blank placeholder for daily unique visitors. Do not copy unique_visitors_14d into this column.",
    "clones_count": "Blank placeholder for daily total clones from GitHub traffic.",
    "clones_uniques": "Blank placeholder for daily unique cloners from GitHub traffic.",
    "collected_at": "UTC timestamp when this screening workbook was generated. Later daily collectors can replace or extend this with actual collection timestamps.",
    "visibility": "Repository visibility copied from the reviewed workbook.",
    "archived": "Repository archived flag copied from the reviewed workbook.",
    "days_since_last_push": "Derived integer number of days between snapshot_date and last_push.",
    "forks_count": "Blank placeholder for repository fork count enrichment.",
    "stargazers_count": "Blank placeholder for repository star count enrichment.",
    "open_issues_count": "Blank placeholder for repository open issues count enrichment.",
}

DAILY_DATABASE_NOTES = [
    (
        "14-day traffic fields",
        (
            "The *_14d values in this screening workbook come from GitHub's rolling "
            "14-day traffic window. They are useful for initial prioritization, but "
            "they should not be the main fact values in the long-term daily database."
        ),
    ),
    (
        "Recommended database shape",
        (
            "Store repository metadata separately from daily traffic facts. Use the "
            "screening workbook to define the repository set, then populate daily "
            "metrics in a dedicated SQLite fact table. The included daily fact columns "
            "are initialized as placeholders because Summary_Reviewed_List.xlsx does "
            "not provide per-day traffic buckets."
        ),
    ),
]

RECOMMENDED_DAILY_DATABASE_FIELDS = [
    (
        "metric_date",
        "The calendar day of the GitHub traffic bucket. This should be the main time key for graphs and heatmaps.",
    ),
    (
        "views_count",
        "Daily total views for the repository on metric_date.",
    ),
    (
        "views_uniques",
        "Daily unique visitors for the repository on metric_date.",
    ),
    (
        "clones_count",
        "Daily total clones for the repository on metric_date.",
    ),
    (
        "clones_uniques",
        "Daily unique cloners for the repository on metric_date.",
    ),
    (
        "collected_at",
        "Timestamp when the data was fetched. Useful for auditability and reruns.",
    ),
]

RECOMMENDED_REPOSITORY_DIMENSION_FIELDS = [
    (
        "repository_full_name",
        "Stable repository key in owner/repository format for joins across all daily records.",
    ),
    (
        "organization",
        "Organization or owner name for grouping and filtering.",
    ),
    (
        "link_2_repository",
        "Canonical repository URL for drill-down navigation.",
    ),
    (
        "last_push",
        "Latest known push timestamp from the repository snapshot.",
    ),
    (
        "traffic_status",
        "Capture restricted or incomplete traffic reads separately from the numeric daily facts.",
    ),
    (
        "forks_count",
        "Repository-level context value that helps compare traffic with reuse and community activity.",
    ),
    (
        "stargazers_count",
        "Slow-moving popularity baseline that complements daily traffic trends.",
    ),
    (
        "open_issues_count",
        "Maintenance-load context that can be compared with traffic attention.",
    ),
    (
        "days_since_last_push",
        "Derived freshness indicator that works well in heatmaps and trend dashboards.",
    ),
]

COLUMN_WIDTHS = {
    "repository": 40,
    "last_push": 24,
    "unique_visitors_14d": 20,
    "unique_cloners_14d": 20,
    "total_views_14d": 18,
    "snapshot_date": 14,
    "organization": 18,
    "repository_full_name": 42,
    "link_2_repository": 65,
    "traffic_status": 18,
    "metric_date": 14,
    "views_count": 14,
    "views_uniques": 14,
    "clones_count": 14,
    "clones_uniques": 14,
    "collected_at": 24,
    "visibility": 14,
    "archived": 12,
    "days_since_last_push": 20,
    "forks_count": 14,
    "stargazers_count": 16,
    "open_issues_count": 18,
}

SHEET_ORGANIZATION_ALIASES = {
    "open-cmisis-pack_raw_list": "open-cmsis-pack",
    "open-cmisis-pack_reviewed_list": "open-cmsis-pack",
}


@dataclass(frozen=True)
class ScreeningRepositoryRow:
    repository: str
    last_push: datetime | str | None
    unique_visitors_14d: int | None
    unique_cloners_14d: int | None
    total_views_14d: int | None
    snapshot_date: date
    organization: str
    repository_full_name: str
    link_2_repository: str
    traffic_status: str
    metric_date: date | None
    views_count: int | None
    views_uniques: int | None
    clones_count: int | None
    clones_uniques: int | None
    collected_at: datetime
    visibility: str
    archived: bool | str | None
    days_since_last_push: int | None
    forks_count: int | None
    stargazers_count: int | None
    open_issues_count: int | None


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read Summary_Reviewed_List.xlsx, keep only ignore_status=USE rows, "
            "and create an .xlsx screening workbook for later SQLite imports."
        )
    )
    parser.add_argument(
        "reviewed_workbook",
        type=Path,
        help="Reviewed workbook path.",
    )
    parser.add_argument(
        "output_workbook",
        nargs="?",
        type=Path,
        help="Target .xlsx file. Default: <reviewed_workbook_stem>_to_screen.xlsx.",
    )
    parser.add_argument(
        "--snapshot-date",
        type=parse_snapshot_date,
        default=datetime.now(UTC).date(),
        help="Snapshot date in YYYY-MM-DD format. Default: today's UTC date.",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s {SCRIPT_VERSION}",
    )
    return parser.parse_args(argv)


def parse_snapshot_date(value: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError as error:
        raise argparse.ArgumentTypeError(
            "snapshot date must use the format YYYY-MM-DD"
        ) from error


def resolve_output_workbook(
    reviewed_workbook: Path,
    output_workbook: Path | None,
) -> Path:
    if output_workbook is not None:
        return output_workbook.expanduser().resolve()
    return get_repository_root() / DEFAULT_OUTPUT_SUBDIR / (
        reviewed_workbook.stem + DEFAULT_OUTPUT_SUFFIX
    )


def get_repository_root() -> Path:
    return Path(__file__).resolve().parent.parent


def normalize_row_values(values: list[Any]) -> list[str]:
    normalized: list[str] = []
    for value in values[: len(EXPECTED_HEADER)]:
        normalized.append("" if value is None else str(value).strip())
    return normalized


def find_header_row(worksheet) -> tuple[int, dict[str, int]]:
    for row_index in range(1, worksheet.max_row + 1):
        row_values = [
            worksheet.cell(row=row_index, column=column_index).value
            for column_index in range(1, len(EXPECTED_HEADER) + 1)
        ]
        if normalize_row_values(row_values) == EXPECTED_HEADER:
            return row_index, {name: index + 1 for index, name in enumerate(EXPECTED_HEADER)}

    raise ValueError(
        f"Could not locate the repository data header row in worksheet '{worksheet.title}'."
    )


def cell_to_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=UTC)
        else:
            value = value.astimezone(UTC)
        return value.strftime("%Y-%m-%dT%H:%M:%SZ")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    return str(value).strip()


def derive_organization_from_sheet(sheet_name: str) -> str:
    if sheet_name in SHEET_ORGANIZATION_ALIASES:
        return SHEET_ORGANIZATION_ALIASES[sheet_name]
    for suffix in ("_raw_list", "_reviewed_list"):
        if sheet_name.endswith(suffix):
            return sheet_name[: -len(suffix)]
    return sheet_name


def extract_repository_url(sheet_name: str, repository: str, link_cell) -> str:
    if link_cell.hyperlink and link_cell.hyperlink.target:
        return link_cell.hyperlink.target.strip()

    value = cell_to_string(link_cell.value)
    if value.startswith("http://") or value.startswith("https://"):
        return value

    if value.startswith("=HYPERLINK("):
        match = re.search(r'"https://github\.com/"\s*&\s*"([^"]+)"', value)
        if match:
            return f"https://github.com/{match.group(1)}/{repository}"

    organization = derive_organization_from_sheet(sheet_name)
    return f"https://github.com/{organization}/{repository}"


def extract_owner_and_repo(repository_url: str) -> tuple[str, str]:
    parsed = urlparse(repository_url)
    parts = [part for part in parsed.path.split("/") if part]
    if len(parts) < 2:
        raise ValueError(f"Unable to parse owner/repository from URL: {repository_url}")
    return parts[0], parts[1]


def parse_datetime_cell(value: Any) -> datetime | str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.astimezone(UTC).replace(tzinfo=None)
        return value
    if not isinstance(value, str):
        return str(value).strip() or None

    normalized = value.strip()
    if not normalized:
        return None

    try:
        parsed = datetime.fromisoformat(normalized.replace("Z", "+00:00"))
    except ValueError:
        return normalized

    if parsed.tzinfo is not None:
        return parsed.astimezone(UTC).replace(tzinfo=None)
    return parsed


def parse_integer_cell(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        normalized = value.strip()
        if not normalized:
            return None
        if re.fullmatch(r"-?\d+(?:\.0+)?", normalized):
            return int(float(normalized))
    return None


def parse_boolean_cell(value: Any) -> bool | str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        normalized = value.strip().upper()
        if not normalized:
            return None
        if normalized == "TRUE":
            return True
        if normalized == "FALSE":
            return False
        return value.strip()
    return str(value).strip() or None


def calculate_days_since_last_push(
    last_push: datetime | str | None,
    snapshot_date: date,
) -> int | None:
    if not isinstance(last_push, datetime):
        return None
    delta_days = (snapshot_date - last_push.date()).days
    return max(delta_days, 0)


def read_use_rows(
    reviewed_workbook: Path,
    snapshot_date: date,
    collected_at: datetime,
) -> list[ScreeningRepositoryRow]:
    try:
        workbook = load_workbook(reviewed_workbook, data_only=False)
    except PermissionError as error:
        raise SystemExit(
            f"Cannot open workbook '{reviewed_workbook}'. Close the file in Excel and retry."
        ) from error

    selected_rows: list[ScreeningRepositoryRow] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_row, column_map = find_header_row(worksheet)

        for row_index in range(header_row + 1, worksheet.max_row + 1):
            repository = cell_to_string(
                worksheet.cell(row=row_index, column=column_map["repository"]).value
            )
            if not repository:
                continue

            ignore_status = cell_to_string(
                worksheet.cell(row=row_index, column=column_map["ignore_status"]).value
            ).upper()
            if ignore_status != USE_STATUS:
                continue

            link_cell = worksheet.cell(row=row_index, column=column_map["link_2_repository"])
            repository_url = extract_repository_url(sheet_name, repository, link_cell)
            organization, repository_name = extract_owner_and_repo(repository_url)
            last_push = parse_datetime_cell(
                worksheet.cell(row=row_index, column=column_map["last_push"]).value
            )

            selected_rows.append(
                ScreeningRepositoryRow(
                    repository=repository_name,
                    last_push=last_push,
                    unique_visitors_14d=parse_integer_cell(
                        worksheet.cell(
                            row=row_index, column=column_map["unique_visitors_14d"]
                        ).value
                    ),
                    unique_cloners_14d=parse_integer_cell(
                        worksheet.cell(
                            row=row_index, column=column_map["unique_cloners_14d"]
                        ).value
                    ),
                    total_views_14d=parse_integer_cell(
                        worksheet.cell(
                            row=row_index, column=column_map["total_views_14d"]
                        ).value
                    ),
                    snapshot_date=snapshot_date,
                    organization=organization,
                    repository_full_name=f"{organization}/{repository_name}",
                    link_2_repository=repository_url,
                    traffic_status=cell_to_string(
                        worksheet.cell(row=row_index, column=column_map["traffic_status"]).value
                    ),
                    metric_date=None,
                    views_count=None,
                    views_uniques=None,
                    clones_count=None,
                    clones_uniques=None,
                    collected_at=collected_at,
                    visibility=cell_to_string(
                        worksheet.cell(row=row_index, column=column_map["visibility"]).value
                    ),
                    archived=parse_boolean_cell(
                        worksheet.cell(row=row_index, column=column_map["archived"]).value
                    ),
                    days_since_last_push=calculate_days_since_last_push(
                        last_push,
                        snapshot_date,
                    ),
                    forks_count=None,
                    stargazers_count=None,
                    open_issues_count=None,
                )
            )

    return selected_rows


def apply_column_widths(worksheet) -> None:
    for field_name, width in COLUMN_WIDTHS.items():
        column_letter = get_column_letter(OUTPUT_COLUMN_INDEX[field_name])
        worksheet.column_dimensions[column_letter].width = width


def apply_table_header_style(worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL


def apply_alternating_row_fill(worksheet, last_row: int) -> None:
    for row_index in range(2, last_row + 1):
        if (row_index - 2) % 2 != 0:
            continue

        for column_index in range(1, len(OUTPUT_COLUMNS) + 1):
            worksheet.cell(row=row_index, column=column_index).fill = ALTERNATING_ROW_FILL


def apply_data_formats(worksheet, first_data_row: int, last_data_row: int) -> None:
    for row_index in range(first_data_row, last_data_row + 1):
        worksheet.cell(
            row=row_index,
            column=OUTPUT_COLUMN_INDEX["last_push"],
        ).number_format = DATE_TIME_NUMBER_FORMAT
        worksheet.cell(
            row=row_index,
            column=OUTPUT_COLUMN_INDEX["snapshot_date"],
        ).number_format = DATE_NUMBER_FORMAT
        worksheet.cell(
            row=row_index,
            column=OUTPUT_COLUMN_INDEX["metric_date"],
        ).number_format = DATE_NUMBER_FORMAT
        worksheet.cell(
            row=row_index,
            column=OUTPUT_COLUMN_INDEX["collected_at"],
        ).number_format = DATE_TIME_NUMBER_FORMAT
        worksheet.cell(
            row=row_index,
            column=OUTPUT_COLUMN_INDEX["link_2_repository"],
        ).font = HYPERLINK_FONT
        for column_index in (
            OUTPUT_COLUMN_INDEX["unique_visitors_14d"],
            OUTPUT_COLUMN_INDEX["unique_cloners_14d"],
            OUTPUT_COLUMN_INDEX["total_views_14d"],
            OUTPUT_COLUMN_INDEX["views_count"],
            OUTPUT_COLUMN_INDEX["views_uniques"],
            OUTPUT_COLUMN_INDEX["clones_count"],
            OUTPUT_COLUMN_INDEX["clones_uniques"],
            OUTPUT_COLUMN_INDEX["days_since_last_push"],
            OUTPUT_COLUMN_INDEX["forks_count"],
            OUTPUT_COLUMN_INDEX["stargazers_count"],
            OUTPUT_COLUMN_INDEX["open_issues_count"],
        ):
            worksheet.cell(row=row_index, column=column_index).number_format = INTEGER_NUMBER_FORMAT


def add_table(worksheet, last_row: int) -> None:
    last_column_letter = get_column_letter(len(OUTPUT_COLUMNS))
    table = Table(displayName=DATA_SHEET_NAME, ref=f"A1:{last_column_letter}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium7",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def write_data_sheet(workbook: Workbook, rows: list[ScreeningRepositoryRow]) -> None:
    worksheet = workbook.create_sheet(DATA_SHEET_NAME)
    worksheet.freeze_panes = "A2"

    for column_index, field_name in enumerate(OUTPUT_COLUMNS, start=1):
        worksheet.cell(row=1, column=column_index, value=field_name)

    for row_index, row in enumerate(rows, start=2):
        for column_index, field_name in enumerate(OUTPUT_COLUMNS, start=1):
            value = getattr(row, field_name)
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            if field_name == "link_2_repository" and value:
                cell.hyperlink = value

    last_row = len(rows) + 1
    add_table(worksheet, last_row)
    apply_column_widths(worksheet)
    apply_table_header_style(worksheet)
    apply_alternating_row_fill(worksheet, last_row)
    apply_data_formats(worksheet, 2, last_row)


def write_readme_sheet(
    workbook: Workbook,
    reviewed_workbook: Path,
    output_workbook: Path,
    snapshot_date: date,
    row_count: int,
) -> None:
    worksheet = workbook.create_sheet(README_SHEET_NAME)
    generated_on = datetime.now(UTC).strftime("%Y-%m-%d")

    readme_rows = [
        ("Revision History", ""),
        ("Author", "OpenAI Codex"),
        ("Date", generated_on),
        (
            "Change",
            (
                "Created the screening workbook from Summary_Reviewed_List.xlsx "
                f"using 04_Get_Repositories_to_screen.py v{SCRIPT_VERSION}"
            ),
        ),
        ("", ""),
        ("File Description", ""),
        (
            "Summary",
            (
                "Rows marked USE were copied from the reviewed workbook into a single "
                "sheet for repository screening and later SQLite imports. The copied "
                "traffic values remain GitHub rolling 14-day snapshot values."
            ),
        ),
        (
            "Generation Process",
            (
                "Two-stage process: stage 1 manual review in Summary_Reviewed_List.xlsx; "
                "stage 2 automated filtering into the screening workbook."
            ),
        ),
        ("Source Workbook", str(reviewed_workbook)),
        ("Source Script", f"04_Get_Repositories_to_screen.py v{SCRIPT_VERSION}"),
        ("Output Workbook", str(output_workbook)),
        ("Data Sheet", DATA_SHEET_NAME),
        ("Rows Exported", row_count),
        ("Snapshot Date", snapshot_date.strftime("%Y-%m-%d")),
        ("", ""),
        ("Field Descriptions", ""),
    ]

    for field_name, description in FIELD_DESCRIPTIONS.items():
        readme_rows.append((field_name, description))

    readme_rows.extend(
        [
            ("", ""),
            ("Daily Database Notes", ""),
        ]
    )
    readme_rows.extend(DAILY_DATABASE_NOTES)
    readme_rows.extend(
        [
            ("", ""),
            ("Included Daily Fact Fields", ""),
        ]
    )
    readme_rows.extend(RECOMMENDED_DAILY_DATABASE_FIELDS)
    readme_rows.extend(
        [
            ("", ""),
            ("Included Repository/Context Fields", ""),
        ]
    )
    readme_rows.extend(RECOMMENDED_REPOSITORY_DIMENSION_FIELDS)

    for row_index, (label, value) in enumerate(readme_rows, start=1):
        worksheet.cell(row=row_index, column=1, value=label)
        worksheet.cell(row=row_index, column=2, value=value)

    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 120


def build_workbook(
    reviewed_workbook: Path,
    output_workbook: Path,
    rows: list[ScreeningRepositoryRow],
    snapshot_date: date,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_data_sheet(workbook, rows)
    write_readme_sheet(
        workbook,
        reviewed_workbook,
        output_workbook,
        snapshot_date,
        len(rows),
    )

    output_workbook.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(output_workbook)
    except PermissionError as error:
        raise SystemExit(
            f"Cannot save workbook '{output_workbook}'. Close the file in Excel and retry."
        ) from error


def print_summary(
    reviewed_workbook: Path,
    output_workbook: Path,
    snapshot_date: date,
    rows: list[ScreeningRepositoryRow],
) -> None:
    print(f"Reviewed workbook: {reviewed_workbook}")
    print(f"Rows selected with ignore_status={USE_STATUS}: {len(rows)}")
    print(f"Snapshot date: {snapshot_date.strftime('%Y-%m-%d')}")
    print(f"Output workbook: {output_workbook}")


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    reviewed_workbook = args.reviewed_workbook.expanduser().resolve()
    if not reviewed_workbook.is_file():
        raise SystemExit(f"Reviewed workbook not found: {reviewed_workbook}")

    output_workbook = resolve_output_workbook(reviewed_workbook, args.output_workbook)
    collected_at = datetime.now(UTC).replace(microsecond=0, tzinfo=None)
    rows = read_use_rows(reviewed_workbook, args.snapshot_date, collected_at)
    if not rows:
        raise SystemExit(
            f"No rows with ignore_status={USE_STATUS} were found in {reviewed_workbook}."
        )

    build_workbook(reviewed_workbook, output_workbook, rows, args.snapshot_date)
    print_summary(reviewed_workbook, output_workbook, args.snapshot_date, rows)
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
