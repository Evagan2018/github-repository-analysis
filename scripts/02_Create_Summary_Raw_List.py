"""Create Summary_Raw_List.xlsx from the four repository CSV exports.

Description:
Read the four ``*_raw_list.csv`` files produced by ``01_Get_Repositories.py`` and
create a ``Summary_Raw_List.xlsx`` workbook with one worksheet per
organization.

Parameters:
- ``input_dir``: directory containing the four raw CSV files.
  Default: ``<repo_root>/data/raw``.
- ``output_xlsx``: target workbook path.
  Default: ``<repo_root>/data/reviewed/Summary_Raw_List.xlsx``.

Calling format:
    python scripts/02_Create_Summary_Raw_List.py
    python scripts/02_Create_Summary_Raw_List.py <input_dir>
    python scripts/02_Create_Summary_Raw_List.py <input_dir> <output_xlsx>

Example:
    python scripts/02_Create_Summary_Raw_List.py

Dependencies:
The script requires ``openpyxl``.
Install it with ``pip install openpyxl`` if it is not already available.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover - handled at runtime
    raise SystemExit(
        "openpyxl is required to generate Summary_Raw_List.xlsx. "
        "Install it with: pip install openpyxl"
    ) from exc


SCRIPT_VERSION = "1.0.2"
DEFAULT_WORKBOOK_NAME = "Summary_Raw_List.xlsx"
DEFAULT_INPUT_SUBDIR = Path("data") / "raw"
DEFAULT_OUTPUT_SUBDIR = Path("data") / "reviewed"
DATE_NUMBER_FORMAT = 'yyyy-mm-dd"T"hh:mm:ss"Z"'
INTEGER_NUMBER_FORMAT = "0"
OUTPUT_COLUMNS = 10
TABLE_HEADER = [f"Column{index}" for index in range(1, OUTPUT_COLUMNS + 1)]
HYPERLINK_FONT = Font(color="0563C1", underline="single")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="70AD47")
ALTERNATING_ROW_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")

DATA_HEADER = [
    "repository",
    "last_push",
    "visibility",
    "archived",
    "unique_visitors_14d",
    "unique_cloners_14d",
    "total_views_14d",
    "traffic_status",
]

COLUMN_WIDTHS = {
    "A": 40,
    "B": 120,
    "C": 13,
    "D": 13,
    "E": 21,
    "F": 21,
    "G": 18,
    "H": 50,
    "I": 60,
    "J": 20,
}


@dataclass(frozen=True)
class WorkbookSource:
    csv_file: str
    sheet_name: str
    github_org: str
    table_name: str


REQUIRED_FILES = [
    WorkbookSource(
        csv_file="open-cmisis-pack_raw_list.csv",
        sheet_name="open-cmisis-pack_raw_list",
        github_org="open-cmsis-pack",
        table_name="open_cmisis_pack_raw_list",
    ),
    WorkbookSource(
        csv_file="arm-examples_raw_list.csv",
        sheet_name="arm-examples_raw_list",
        github_org="arm-examples",
        table_name="arm_examples_raw_list",
    ),
    WorkbookSource(
        csv_file="arm-software_raw_list.csv",
        sheet_name="arm-software_raw_list",
        github_org="arm-software",
        table_name="arm_software_raw_list",
    ),
    WorkbookSource(
        csv_file="mdk-packs_raw_list.csv",
        sheet_name="mdk-packs_raw_list",
        github_org="mdk-packs",
        table_name="mdk_packs_raw_list",
    ),
]


def parse_args(argv: list[str]) -> tuple[Path, Path]:
    parser = argparse.ArgumentParser(
        description=(
            "Read the four *_raw_list.csv files produced by 01_Get_Repositories.py "
            "and create a Summary_Raw_List.xlsx workbook."
        )
    )
    parser.add_argument(
        "input_dir",
        nargs="?",
        help="Directory containing the four raw CSV files.",
    )
    parser.add_argument(
        "output_xlsx",
        nargs="?",
        help="Output workbook path. Default: <input_dir>/Summary_Raw_List.xlsx.",
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s {SCRIPT_VERSION}",
    )
    args = parser.parse_args(argv)

    repo_root = get_repository_root()
    default_input_dir = repo_root / DEFAULT_INPUT_SUBDIR
    default_output_xlsx = repo_root / DEFAULT_OUTPUT_SUBDIR / DEFAULT_WORKBOOK_NAME
    input_dir = (
        Path(args.input_dir).expanduser().resolve()
        if args.input_dir
        else default_input_dir
    )
    output_xlsx = (
        Path(args.output_xlsx).expanduser().resolve()
        if args.output_xlsx
        else default_output_xlsx
    )
    return input_dir, output_xlsx


def get_repository_root() -> Path:
    return Path(__file__).resolve().parent.parent


def read_csv_rows(file_path: Path) -> list[list[str]]:
    with file_path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.reader(handle))


def normalize_cell(value: str) -> str | None:
    return None if value == "" else value


def parse_datetime_cell(value: str | None) -> datetime | str | None:
    if not isinstance(value, str):
        return value

    normalized = value.strip()
    if not normalized:
        return None

    try:
        parsed = datetime.fromisoformat(normalized.replace("Z", "+00:00"))
    except ValueError:
        return value

    if parsed.tzinfo is not None:
        parsed = parsed.astimezone(UTC).replace(tzinfo=None)

    return parsed


def parse_boolean_cell(value: str | None) -> bool | str | None:
    if not isinstance(value, str):
        return value

    normalized = value.strip().upper()
    if normalized == "TRUE":
        return True
    if normalized == "FALSE":
        return False
    return value


def parse_numeric_cell(value: str | None) -> int | float | str | None:
    if not isinstance(value, str):
        return value

    normalized = value.strip()
    if not normalized:
        return None

    if not re.fullmatch(r"-?\d+(?:\.\d+)?", normalized):
        return value

    if "." in normalized:
        return float(normalized)
    return int(normalized)


def pad_row(row: Iterable[str | None], width: int) -> list[str | None]:
    padded = list(row)[:width]
    while len(padded) < width:
        padded.append(None)
    return padded


def find_data_header_index(rows: list[list[str]], file_name: str) -> int:
    for index, row in enumerate(rows):
        if all((row[column] if column < len(row) else "") == header for column, header in enumerate(DATA_HEADER)):
            return index
    raise ValueError(f"Could not locate the repository data header row in {file_name}.")


def coerce_repository_row(row: list[object]) -> None:
    row[1] = parse_datetime_cell(row[1])
    row[3] = parse_boolean_cell(row[3])
    row[4] = parse_numeric_cell(row[4])
    row[5] = parse_numeric_cell(row[5])
    row[6] = parse_numeric_cell(row[6])


def build_sheet_rows(
    csv_rows: list[list[str]],
    github_org: str,
    file_name: str,
) -> tuple[list[list[object]], int | None, int | None]:
    header_index = find_data_header_index(csv_rows, file_name)
    sheet_rows: list[list[object]] = [TABLE_HEADER.copy()]
    first_repository_excel_row: int | None = None
    last_repository_excel_row: int | None = None

    for csv_row_index, csv_row in enumerate(csv_rows):
        row: list[object] = pad_row((normalize_cell(value) for value in csv_row), OUTPUT_COLUMNS)

        if csv_row_index == header_index:
            row[8] = "link_2_repository"
            row[9] = "ignore_status"
        elif csv_row_index > header_index and row[0]:
            coerce_repository_row(row)
            excel_row = csv_row_index + 2
            row[8] = (
                f'=HYPERLINK("https://github.com/" & "{github_org}" & "/" & A{excel_row}, A{excel_row})'
            )
            row[9] = None
            if first_repository_excel_row is None:
                first_repository_excel_row = excel_row
            last_repository_excel_row = excel_row
        else:
            row[8] = row[8] if row[8] not in ("", None) else None
            row[9] = row[9] if row[9] not in ("", None) else None

        sheet_rows.append(row)

    return sheet_rows, first_repository_excel_row, last_repository_excel_row


def apply_column_widths(worksheet) -> None:
    for column, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column].width = width


def apply_table_header_style(worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL


def apply_alternating_row_fill(worksheet, last_row: int) -> None:
    for row_index in range(2, last_row + 1):
        if (row_index - 2) % 2 != 0:
            continue

        for column_index in range(1, OUTPUT_COLUMNS + 1):
            worksheet.cell(row=row_index, column=column_index).fill = ALTERNATING_ROW_FILL


def apply_repository_formats(
    worksheet,
    first_repository_excel_row: int | None,
    last_repository_excel_row: int | None,
) -> None:
    if first_repository_excel_row is None or last_repository_excel_row is None:
        return

    for row_index in range(first_repository_excel_row, last_repository_excel_row + 1):
        worksheet.cell(row=row_index, column=2).number_format = DATE_NUMBER_FORMAT
        worksheet.cell(row=row_index, column=9).font = HYPERLINK_FONT
        for column_index in (5, 6, 7):
            worksheet.cell(row=row_index, column=column_index).number_format = INTEGER_NUMBER_FORMAT


def write_rows(worksheet, rows: list[list[object]]) -> None:
    for row_index, row_values in enumerate(rows, start=1):
        for column_index, value in enumerate(row_values, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)


def add_table(worksheet, table_name: str, last_row: int) -> None:
    table = Table(displayName=table_name, ref=f"A1:J{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium7",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def build_workbook(input_dir: Path, output_xlsx: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for source in REQUIRED_FILES:
        csv_path = input_dir / source.csv_file
        csv_rows = read_csv_rows(csv_path)
        sheet_rows, first_repository_excel_row, last_repository_excel_row = build_sheet_rows(
            csv_rows,
            source.github_org,
            source.csv_file,
        )

        worksheet = workbook.create_sheet(source.sheet_name)
        write_rows(worksheet, sheet_rows)
        apply_column_widths(worksheet)
        add_table(worksheet, source.table_name, len(sheet_rows))
        apply_table_header_style(worksheet)
        apply_alternating_row_fill(worksheet, len(sheet_rows))
        apply_repository_formats(
            worksheet,
            first_repository_excel_row,
            last_repository_excel_row,
        )

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)


def validate_input_files(input_dir: Path) -> None:
    for source in REQUIRED_FILES:
        file_path = input_dir / source.csv_file
        if not file_path.is_file():
            raise FileNotFoundError(f"Missing input file: {file_path}")


def main(argv: list[str]) -> int:
    input_dir, output_xlsx = parse_args(argv)
    validate_input_files(input_dir)
    build_workbook(input_dir, output_xlsx)
    print(f"Created workbook: {output_xlsx}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
